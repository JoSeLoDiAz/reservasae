# -*- coding: utf-8 -*-
"""Extrae el catálogo de los proyectos a catalogo.json."""

import json
import math
import os
import unicodedata

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SALIDA = os.path.join(RAIZ, "backend", "prisma", "seed", "catalogo.json")

SOBRECUPO = 0.30

PROYECTOS = [
    {
        "slug": "britcham-adee",
        "archivo": os.path.join(RAIZ, "docs", "proyectos", "06 - BRITCHAM - ADEE.xlsx"),
    },
    {
        "slug": "adecopria",
        "archivo": os.path.join(RAIZ, "docs", "proyectos", "01 - ADECOPRIA.xlsx"),
    },
]


def limpiar(valor):
    """Quita espacios y saltos sobrantes."""
    if valor is None:
        return None
    texto = str(valor).replace("_x000D_", " ").replace("\r", " ").replace("\n", " ")
    texto = " ".join(texto.split()).strip()
    return texto or None


def normalizar_ubicacion(valor):
    """Normaliza el nombre de la ubicación."""
    texto = limpiar(valor)
    if texto is None:
        return None
    return texto.upper()


def clave_sin_tildes(texto):
    """Quita las tildes para comparar."""
    descompuesto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in descompuesto if unicodedata.category(c) != "Mn")


def repartir_sobrecupo(celdas):
    """Reparte el 30 % de sobrecupo de un grupo."""
    total_base = sum(c["cuposBase"] for c in celdas)
    # floor(x + 0,5): round() redondea al par
    total_maximo = int(math.floor(total_base * (1 + SOBRECUPO) + 0.5))

    exactos = [c["cuposBase"] * (1 + SOBRECUPO) for c in celdas]
    asignados = [int(math.floor(v)) for v in exactos]
    faltan = total_maximo - sum(asignados)

    orden = sorted(
        range(len(celdas)),
        key=lambda i: (
            -(exactos[i] - asignados[i]),
            -celdas[i]["cuposBase"],
            celdas[i]["ubicacion"],
        ),
    )
    for i in orden[:faltan]:
        asignados[i] += 1

    for celda, maximo in zip(celdas, asignados):
        celda["cuposMaximos"] = maximo


def columna(encabezados, *fragmentos):
    """Busca una columna por fragmento."""
    for indice, texto in enumerate(encabezados):
        arriba = clave_sin_tildes(texto.upper())
        if all(clave_sin_tildes(f.upper()) in arriba for f in fragmentos):
            return indice
    return None


def leer_datos_basicos(libro):
    hoja = libro["Datos_Basicos"]
    filas = list(hoja.iter_rows(values_only=True))
    encabezados = [str(h) if h is not None else "" for h in filas[0]]
    fila = filas[1]

    def dato(*fragmentos):
        indice = columna(encabezados, *fragmentos)
        return limpiar(fila[indice]) if indice is not None else None

    return {
        "nombre": dato("NOMBRE DE LA ENTIDAD"),
        "sigla": dato("SIGLA"),
        "nit": dato("NUMERO DE IDENTIFICACION") or dato("NÚMERO DE IDENTIFICACION"),
        "digitoVerificacion": dato("DIGITO VERIFICACION"),
        "correo": dato("CORREO"),
        "telefono": dato("TELEFONO"),
        "departamento": dato("DEPARTAMENTO DE DOMICILIO"),
        "ciudad": dato("CIUDAD/MUNICIPIO DE DOMICILIO"),
        "direccion": dato("DIRECCION DE DOMICILIO"),
        "paginaWeb": dato("PAGINA WEB"),
    }


def leer_acciones(libro):
    hoja = libro["Datos_AF"]
    filas = list(hoja.iter_rows(values_only=True))
    encabezados = [str(h) if h is not None else "" for h in filas[0]]

    indices = {
        "consecutivo": columna(encabezados, "CONSECUTIVO DE LA ACCION"),
        "nombre": columna(encabezados, "NOMBRE DE LA ACCION"),
        "modalidad": columna(encabezados, "MODALIDAD DE FORMACION"),
        "evento": columna(encabezados, "EVENTO DE FORMACION"),
        "enfoque": columna(encabezados, "ENFOQUE DE LA ACCION"),
        "metodologia": columna(encabezados, "METODOLOGIA DE FORMACION"),
        "horas": columna(encabezados, "NUMERO DE HORAS POR GRUPO"),
        "objetivo": columna(encabezados, "OBJETIVO"),
        "ambiente": columna(encabezados, "AMBIENTE DE APRENDIZAJE"),
    }

    acciones = {}
    for orden, fila in enumerate(filas[1:], start=1):
        if indices["consecutivo"] is None or fila[indices["consecutivo"]] is None:
            continue
        consecutivo = int(fila[indices["consecutivo"]])
        acciones[consecutivo] = {
            "codigo": "AF%d" % consecutivo,
            "consecutivo": consecutivo,
            "orden": orden,
            "nombre": limpiar(fila[indices["nombre"]]),
            "modalidad": (limpiar(fila[indices["modalidad"]]) or "").upper() or None,
            "evento": limpiar(fila[indices["evento"]]),
            "enfoque": limpiar(fila[indices["enfoque"]]),
            "metodologia": limpiar(fila[indices["metodologia"]]),
            "horas": int(fila[indices["horas"]]) if fila[indices["horas"]] else None,
            "objetivo": limpiar(fila[indices["objetivo"]]),
            "ambiente": limpiar(fila[indices["ambiente"]]),
            "grupos": [],
        }
    return acciones


def leer_cobertura(libro, acciones):
    hoja = libro["Datos_Cobertura"]
    filas = list(hoja.iter_rows(values_only=True))

    for fila in filas[1:]:
        if fila[0] is None:
            continue
        consecutivo = int(fila[0])
        accion = acciones.get(consecutivo)
        if accion is None:
            continue

        # "GRUPO 3" -> 3, el identificador del grupo
        etiqueta = limpiar(fila[1]) or ""
        digitos = "".join(c for c in etiqueta if c.isdigit())
        numero = int(digitos) if digitos else len(accion["grupos"]) + 1

        coberturas = []

        # sede presencial
        ciudad = normalizar_ubicacion(fila[3])
        if ciudad and fila[4]:
            coberturas.append(
                {
                    "ubicacion": ciudad,
                    "tipo": "CIUDAD",
                    "departamento": normalizar_ubicacion(fila[2]),
                    "modalidad": "PRESENCIAL",
                    "cuposBase": int(fila[4]),
                }
            )

        # cobertura virtual, hasta 25 pares
        for i in range(5, 55, 2):
            if i + 1 >= len(fila):
                break
            if fila[i] and fila[i + 1]:
                coberturas.append(
                    {
                        "ubicacion": normalizar_ubicacion(fila[i]),
                        "tipo": "DEPARTAMENTO",
                        "departamento": normalizar_ubicacion(fila[i]),
                        "modalidad": "VIRTUAL",
                        "cuposBase": int(fila[i + 1]),
                    }
                )

        repartir_sobrecupo(coberturas)

        sede = next((c for c in coberturas if c["modalidad"] == "PRESENCIAL"), None)
        accion["grupos"].append(
            {
                "numero": numero,
                "sedeCiudad": sede["ubicacion"] if sede else None,
                "sedeDepartamento": sede["departamento"] if sede else None,
                "coberturas": coberturas,
            }
        )

    for accion in acciones.values():
        accion["grupos"].sort(key=lambda g: g["numero"])


def construir_ofertas(accion):
    """Suma los cupos de los grupos por ubicación."""
    acumulado = {}
    for grupo in accion["grupos"]:
        for cobertura in grupo["coberturas"]:
            clave = (cobertura["ubicacion"], cobertura["tipo"])
            if clave not in acumulado:
                acumulado[clave] = {
                    "ubicacion": cobertura["ubicacion"],
                    "tipo": cobertura["tipo"],
                    "departamento": cobertura["departamento"],
                    "modalidad": cobertura["modalidad"],
                    "cuposBase": 0,
                    "cuposMaximos": 0,
                    "grupos": [],
                }
            acumulado[clave]["cuposBase"] += cobertura["cuposBase"]
            acumulado[clave]["cuposMaximos"] += cobertura["cuposMaximos"]
            acumulado[clave]["grupos"].append(grupo["numero"])
    return sorted(acumulado.values(), key=lambda o: (o["tipo"], o["ubicacion"]))


def main():
    convenios = []
    for proyecto in PROYECTOS:
        libro = openpyxl.load_workbook(proyecto["archivo"], data_only=True)
        basicos = leer_datos_basicos(libro)
        acciones = leer_acciones(libro)
        leer_cobertura(libro, acciones)

        lista = []
        for accion in sorted(acciones.values(), key=lambda a: a["consecutivo"]):
            accion["ofertas"] = construir_ofertas(accion)
            lista.append(accion)

        convenios.append(
            {
                "slug": proyecto["slug"],
                "entidad": basicos,
                "archivoOrigen": os.path.basename(proyecto["archivo"]),
                "acciones": lista,
            }
        )

    catalogo = {
        "sobrecupo": SOBRECUPO,
        "nota": (
            "Generado por scripts/extraer-catalogo.py desde docs/proyectos/. "
            "cuposMaximos ya incluye el 30% de sobrecupo por deserción."
        ),
        "convenios": convenios,
    }

    os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
    with open(SALIDA, "w", encoding="utf-8") as destino:
        json.dump(catalogo, destino, ensure_ascii=False, indent=2)

    for convenio in convenios:
        grupos = sum(len(a["grupos"]) for a in convenio["acciones"])
        ofertas = sum(len(a["ofertas"]) for a in convenio["acciones"])
        base = sum(o["cuposBase"] for a in convenio["acciones"] for o in a["ofertas"])
        maximo = sum(o["cuposMaximos"] for a in convenio["acciones"] for o in a["ofertas"])
        print(
            "%-16s acciones=%-3d grupos=%-3d ofertas=%-3d base=%-5d maximo=%d"
            % (convenio["slug"], len(convenio["acciones"]), grupos, ofertas, base, maximo)
        )
    print("escrito en %s" % SALIDA)


if __name__ == "__main__":
    main()
